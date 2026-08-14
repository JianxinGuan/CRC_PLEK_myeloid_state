from openpyxl import load_workbook
from pathlib import Path

path = Path(r"G:\New_CRC_Platelet\24_制作投稿文件包\supplementary\Supplementary_Tables_S1-S8.xlsx")
wb = load_workbook(path)
ws = wb["Table2_Evidence"]
rows = {
2: ("The frozen PF4/PPBP/PLEK composite transcriptional axis was consistently associated with a neutrophil-related inflammatory tissue state in CRC bulk tissue.", "Five GEO cohorts, n=1,055; pooled rho=0.400, 95% CI 0.287-0.502, FDR=1.86e-10; concordant direction in 5/5 cohorts; prediction interval 0.159-0.596.", "This is a tissue-state association and does not establish cellular origin, direct contact, or causality."),
3: ("The composite association had unequal components: PLEK was the principal contributor, PPBP was a stable secondary bulk component, and PF4 showed no independent association.", "PLEK rho=0.693; PPBP rho=0.290; PF4 rho=-0.004. Excluding PF4 gave rho=0.509; matched-random-module empirical P=0.000999.", "The frozen three-gene score was retained for transparency; the three genes should not be described as equivalent drivers."),
4: ("PLEK is a credible endogenous myeloid/white-cell signal enriched in neutrophils but not neutrophil-specific; between-sample bulk covariance primarily reflects differences in PLEK-rich myeloid/neutrophil composition.", "Tumor tissue: 1,446,812 cells and 381 donors; PLEK detection in neutrophils 86.6%; study-adjusted donor bulk-like beta=0.542.", "PLEK should not be described as a neutrophil-specific marker."),
5: ("PPBP/PF4 signals in neutrophils were sparse and study-dependent, compatible with ambient RNA, residual blood, or platelet-leukocyte complexes.", "Neutrophil detection: PPBP 0.39%, PF4 0.16%; each covered eight donors; study-adjusted donor associations were not robust.", "The data neither prove neutrophil transcription nor exclude true platelet contact."),
6: ("PLEK and the three-gene axis showed weak average spatial associations in CRC, with high between-patient heterogeneity and no neutrophil specificity.", "Seven patients and 14 sections; PLEK-neutrophil rho=0.081, I2=86.1%; three-gene-axis rho=0.066, I2=89.7%; prediction intervals crossed zero; PLEK-endothelial rho=0.079.", "The data do not support claims of stable colocalization, cell contact, platelet origin, or spatial mechanism."),
7: ("The frozen 13-gene platelet-related state had no CRC overall-survival association.", "Five GEO cohorts, n=1,044, 398 events; HR=1.006, 95% CI 0.911-1.112, P=0.903; stage-adjusted HR=1.000; matched-random-module empirical P=0.927.", "Do not develop or describe this state as a prognostic model, and do not select cut-points, cohorts, or genes post hoc to optimize survival results."),
8: ("The positive association between the 10-gene identity/adhesion axis and the endothelial module did not replicate across GEO cohorts.", "Five-cohort pooled rho=-0.157, 95% CI -0.334 to 0.031, FDR=0.102, I2=81.9%; only 1/5 cohort directions was positive.", "The positive TCGA association should be described as platform- or cohort-dependent, not as a stable endothelial marker."),
9: ("The frozen transcriptional state showed no stable association with MSI or stage; CMS was not tested because no prespecified reproducible labels were available.", "All TCGA MSI analyses had FDR=0.734; stage analyses had FDR approximately 0.728.", "Do not substitute another PanCancer subtype for CMS or describe an untested analysis as negative."),
}
for r, vals in rows.items():
    for col, value in zip((4,5,6), vals):
        ws.cell(r, col).value = value
ws = wb["TableS1_Boundaries"]
vals = {
2: ("The 13-gene state predicts CRC survival or can support a clinical prognostic model", "Stage 13 produced a robust null result."),
3: ("PF4, PPBP, and PLEK jointly and equally drive the neutrophil association", "PLEK was the principal contributor, PPBP secondary, and PF4 null."),
4: ("The frozen three-gene axis measures intratumoral platelet abundance or is platelet-specific", "Single-cell data indicate predominantly myeloid PLEK expression; PPBP/PF4 origin remains uncertain."),
5: ("PPBP/PF4 are stably transcribed by CRC neutrophils", "Single-cell detection was very low and compatible with ambient RNA or complexes."),
6: ("Spatial analysis validates stable neutrophil-specific colocalization", "Spatial effects were small and heterogeneous; prediction intervals crossed zero and endothelial associations were similar."),
7: ("PF4 has no spatial relationship with platelets", "The spatial data showed no evidence; they cannot prove biological absence."),
8: ("The 10-gene axis is a stable endothelial-background marker", "The direction did not replicate in five GEO cohorts."),
9: ("Correlation proves direct platelet-neutrophil contact, mechanism, or causality", "All available transcriptomic analyses were observational and cannot support causal inference."),
10: ("CMS analysis was negative", "CMS was not run because no prespecified reproducible labels or classifier were available."),
}
for r, (claim, reason) in vals.items():
    ws.cell(r,2).value = claim
    ws.cell(r,3).value = reason
wb.save(path)
print(path)
